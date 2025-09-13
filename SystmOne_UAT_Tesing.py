import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from datetime import datetime

# NHS Blue colour + fills
nhs_blue_fill = PatternFill(start_color="005EB8", end_color="005EB8", fill_type="solid")
alt_row_fill = PatternFill(start_color="E6F2FF", end_color="E6F2FF", fill_type="solid")
pass_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
fail_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
blue_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")

# Borders
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

# Autosize columns with fixed widths
def autosize_columns(ws):
    for col in ws.columns:
        max_length = 0
        col_letter = None
        for cell in col:
            if cell.coordinate in ws.merged_cells:
                continue
            if col_letter is None:
                col_letter = cell.column_letter
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        if col_letter:
            width = max(max_length + 2, 15)
            ws.column_dimensions[col_letter].width = width

    # Fixed widths for important columns
    ws.column_dimensions["A"].width = 70  # Test Step
    ws.column_dimensions["B"].width = 20   # Expected Result
    ws.column_dimensions["C"].width = 20   # Actual Result
    for col in ["E", "F", "G", "H"]:
        ws.column_dimensions[col].width = 15  # merged Comments ~60 total

# Create UAT workbook
def create_uat_workbook(service_name, start_date):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    areas = [
        "Registration and Referral in",
        "Referral Allocation",
        "Triage",
        "Appointments",
        "Rota Templates",
        "Tasks",
        "Letters",
        "Consent & Confidentiality",
        "Communication Annexe",
        "Event Details",
        "Data Templates",
        "Questionnaires"
    ]

    for area in areas:
        ws = wb.create_sheet(title=area[:31])

        # === Row 1 - Title ===
        ws.merge_cells("A1:H1")
        ws["A1"] = f"{service_name} - {area}"
        ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
        ws["A1"].fill = nhs_blue_fill
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

        # === Row 2 - Testing Started ===
        ws["A2"] = "Testing Started:"
        ws["B2"] = start_date
        for col in ["A2", "B2", "C2", "D2"]:
            ws[col].font = Font(bold=True, color="FFFFFF")
            ws[col].fill = nhs_blue_fill
            ws[col].alignment = Alignment(horizontal="left", vertical="center")
        ws["B2"].number_format = "DD/MM/YYYY"
        ws.merge_cells("E2:H2")
        ws["E2"].fill = nhs_blue_fill
        ws["E2"].font = Font(bold=True, color="FFFFFF")

        # === Row 3 - Testing Completed ===
        ws["A3"] = "Testing Completed:"
        ws["B3"] = None
        for col in ["A3", "B3", "C3", "D3"]:
            ws[col].font = Font(bold=True, color="FFFFFF")
            ws[col].fill = nhs_blue_fill
            ws[col].alignment = Alignment(horizontal="left", vertical="center")
        ws["B3"].number_format = "DD/MM/YYYY"
        ws.merge_cells("E3:H3")
        ws["E3"].fill = nhs_blue_fill
        ws["E3"].font = Font(bold=True, color="FFFFFF")

        # === Row 4 - Spacer ===
        for col in range(1, 9):
            cell = ws.cell(row=4, column=col)
            cell.fill = nhs_blue_fill
            cell.font = Font(bold=True, color="FFFFFF")
        ws.merge_cells("E4:H4")

        # === Row 5 - Headers ===
        headers = ["Test Step", "Expected Result", "Actual Result", "Results", "Comments"]
        ws.append(headers)

        # A–D headers
        for col in range(1, 5):
            cell = ws.cell(row=5, column=col)
            cell.fill = nhs_blue_fill
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        # Comments header (E–H merged)
        ws.merge_cells("E5:H5")
        ws["E5"] = "Comments"
        ws["E5"].fill = nhs_blue_fill
        ws["E5"].font = Font(bold=True, color="FFFFFF")
        ws["E5"].alignment = Alignment(horizontal="center", vertical="center")
        ws["E5"].border = thin_border

        # Freeze panes
        ws.freeze_panes = "A6"

        # Data validation for Results column
        dv = DataValidation(type="list", formula1='"Pass,Fail"', allow_blank=True)
        ws.add_data_validation(dv)
        dv.add("D6:D100")

        # Conditional formatting
        ws.conditional_formatting.add("D6:D100",
            CellIsRule(operator="equal", formula=['"Pass"'], fill=pass_fill))
        ws.conditional_formatting.add("D6:D100",
            CellIsRule(operator="equal", formula=['"Fail"'], fill=fail_fill))
        ws.conditional_formatting.add("D6:D100",
            CellIsRule(operator="equal", formula=['""'], fill=blue_fill))

        # Apply formatting + merge Comments column for rows 6–100
        for i in range(6, 101):
            ws.merge_cells(start_row=i, start_column=5, end_row=i, end_column=8)
            for row in ws.iter_rows(min_row=i, max_row=i, min_col=1, max_col=8):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                    cell.border = thin_border
                    if i % 2 == 0:
                        cell.fill = alt_row_fill

        # Autosize columns with fixed widths
        autosize_columns(ws)

    # Save
    file_name = f"UAT_Testing_{service_name}.xlsx"
    wb.save(file_name)
    print(f"Workbook saved as {file_name}")

# === Run Script ===
if __name__ == "__main__":
    service = input("Enter the service name: ")
    start_date_input = input("Enter the Testing Start Date (DD/MM/YYYY): ")
    start_date = datetime.strptime(start_date_input, "%d/%m/%Y").date()
    create_uat_workbook(service, start_date)
